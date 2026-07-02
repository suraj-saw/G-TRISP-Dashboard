# backend/app/routes/surat_export.py
"""
Data export endpoint for the Surat dashboard.
Supports CSV and Excel (.xlsx) downloads, respecting all active filters.

GET /api/surat/dashboard/export
  ?format=csv|excel
  &police_station=...
  &year=...
  &severity=...
  &road_classification=...
  &weather_condition=...
  &light_condition=...
  &collision_type=...
"""

import csv
import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.surat_accident import SuratAccident
from app.utils.surat_accident_utils import apply_surat_filters
from app.utils.text_utils import safe_text
from app.core.constants import SURAT_DASH_PREFIX

router = APIRouter(
    prefix=SURAT_DASH_PREFIX,
    tags=["Surat Dashboard"],
)

# ---------------------------------------------------------------------------
# Column definitions — maps (CSV header, model attribute or derived value)
# ---------------------------------------------------------------------------

EXPORT_COLUMNS = [
    ("Accident ID",               "accident_id"),
    ("District",                  "district"),
    ("Police Station",            "police_station"),
    ("Accident Date Time",        "accident_date_time"),
    ("Latitude",                  "latitude"),
    ("Longitude",                 "longitude"),
    ("Road Name",                 "road_name"),
    ("Road Classification",       "road_classification"),
    ("Severity",                  "severity"),
    ("No of Vehicles",            "number_of_vehicles"),
    # Driver
    ("Drivers Killed",            "driver_killed"),
    ("Drivers Grievous Injury",   "driver_grievous_injury"),
    ("Drivers Minor Injury",      "driver_minor_injury"),
    # Passenger
    ("Passengers Killed",         "passenger_killed"),
    ("Passengers Grievous Injury","passenger_grievous_injury"),
    ("Passengers Minor Injury",   "passenger_minor_injury"),
    # Pedestrian
    ("Pedestrians Killed",        "pedestrian_killed"),
    ("Pedestrians Grievous Injury","pedestrian_grievous_injury"),
    ("Pedestrians Minor Injury",  "pedestrian_minor_injury"),
    # Collision
    ("Collision Type",            "type_of_collision"),
    ("Collision Nature",          "collision_feature"),
    # Conditions
    ("Weather Condition",         "weather_condition"),
    ("Light Condition",           "light_condition"),
    ("Visibility",                "visibility"),
    ("Traffic Violation",         "traffic_violation"),
]

HEADERS = [col[0] for col in EXPORT_COLUMNS]
FIELDS  = [col[1] for col in EXPORT_COLUMNS]


def _row_values(accident: SuratAccident) -> list:
    """Convert a SuratAccident ORM object into an ordered list of cell values."""
    values = []
    for field in FIELDS:
        raw = getattr(accident, field, None)
        if raw is None:
            values.append("")
        elif isinstance(raw, datetime):
            values.append(raw.strftime("%d-%b-%Y %I:%M %p"))
        elif isinstance(raw, float):
            # Preserve None-style floats; round real coords to 6 dp
            values.append(round(raw, 6))
        else:
            # safe_text converts 'nan' / empty to ""
            cleaned = safe_text(str(raw))
            values.append("" if cleaned == "Unknown" else cleaned)
    return values


def _build_filename(fmt: str, filters: dict) -> str:
    parts = ["surat_accidents"]
    if filters.get("police_station"):
        parts.append(filters["police_station"].replace(" ", "_"))
    if filters.get("year"):
        parts.append(str(filters["year"]))
    if filters.get("severity"):
        parts.append(filters["severity"].replace(" ", "_"))
    parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
    ext = "csv" if fmt == "csv" else "xlsx"
    return f"{'_'.join(parts)}.{ext}"


# ---------------------------------------------------------------------------
# Export endpoint
# ---------------------------------------------------------------------------

@router.get("/export")
def export_data(
    format: str = Query("csv", enum=["csv", "excel"]),
    police_station: Optional[List[str]] = Query(None),
    year: Optional[List[int]] = Query(None),
    severity: Optional[List[str]] = Query(None),
    road_classification: Optional[List[str]] = Query(None),
    weather_condition: Optional[List[str]] = Query(None),
    light_condition: Optional[List[str]] = Query(None),
    collision_type: Optional[List[str]] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    # Build query with filters
    query = apply_surat_filters(
        db.query(SuratAccident),
        police_station, year, road_classification,
        weather_condition, light_condition, collision_type,
        date_from, date_to,
    )
    if severity and "all" not in severity:
        query = query.filter(SuratAccident.severity.in_(severity))

    accidents = query.order_by(SuratAccident.accident_date_time).all()

    active_filters = {
        "police_station": police_station,
        "year": year,
        "severity": severity,
        "date_from": date_from,
        "date_to": date_to,
    }
    filename = _build_filename(format, active_filters)

    # ── CSV ──────────────────────────────────────────────────────────────────
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        for acc in accidents:
            writer.writerow(_row_values(acc))

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

    # ── Excel ────────────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl somehow unavailable
        return export_data.__wrapped__(
            format="csv",
            police_station=police_station,
            year=year,
            severity=severity,
            road_classification=road_classification,
            weather_condition=weather_condition,
            light_condition=light_condition,
            collision_type=collision_type,
            date_from=date_from,
            date_to=date_to,
            db=db,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accident Data"

    # ── Header styling ────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="1E3A8A")  # deep blue
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side    = Side(style="thin", color="D1D5DB")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    ws.append(HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border

    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="F8FAFC")  # very light slate

    for row_idx, acc in enumerate(accidents, start=2):
        row_data = _row_values(acc)
        ws.append(row_data)
        row = ws[row_idx]

        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in row:
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.row_dimensions[row_idx].height = 18 if accidents else 18

    # ── Column widths ─────────────────────────────────────────────────────────
    WIDTHS = {
        "Accident ID": 18, "District": 14, "Police Station": 20,
        "Accident Date Time": 22, "Latitude": 12, "Longitude": 12,
        "Road Name": 22, "Road Classification": 20, "Severity": 18,
        "No of Vehicles": 12,
    }
    for col_idx, header in enumerate(HEADERS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = WIDTHS.get(header, 14)

    # ── Freeze header row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Metadata sheet ────────────────────────────────────────────────────────
    meta = wb.create_sheet("Export Info")
    meta_font = Font(bold=True, size=10)
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 36

    meta_rows = [
        ("Export Date", datetime.now().strftime("%d %b %Y %H:%M")),
        ("Total Records", len(accidents)),
        ("Police Station", police_station or "All"),
        ("Year", str(year) if year else "All"),
        ("Severity", severity or "All"),
        ("Road Classification", road_classification or "All"),
        ("Weather Condition", weather_condition or "All"),
        ("Light Condition", light_condition or "All"),
        ("Collision Type", collision_type or "All"),
        ("Source", "G-TRISP Dashboard — Surat Accident Data"),
    ]
    for r_idx, (label, value) in enumerate(meta_rows, start=1):
        meta.cell(r_idx, 1, label).font = meta_font
        meta.cell(r_idx, 2, str(value))

    # ── Stream ────────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
