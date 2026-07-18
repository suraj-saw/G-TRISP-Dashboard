# backend/app/utils/export_utils.py
"""
Data Export Utility Module

Provides shared helper functions for generating downloadable CSV and Excel 
files of accident records. These functions are utilized by the blackspot-export 
endpoints across both the Gujarat and Surat dashboards to ensure consistent 
reporting formats, robust data cleaning, and professional styling.
"""

import csv
import io
from datetime import datetime

from app.utils.text_utils import safe_text

# Column definitions mapping the human-readable output header to the internal ORM attribute.
# The "__bs_id__" field is a special sentinel injected at runtime.
EXPORT_COLUMNS = [
    ("Blackspot #",                "__bs_id__"),
    ("Accident ID",               "accident_id"),
    ("District",                  "district"),
    ("Police Station",            "police_station"),
    ("Accident Date Time",        "accident_date_time"),
    ("Latitude",                  "latitude"),
    ("Longitude",                 "longitude"),
    ("Road Name",                 "road_name"),
    ("Road Classification",       "road_classification"),
    ("Severity",                  "severity"),
    ("No of Vehicles",            "number_of_vehicles"),
    # Driver
    ("Drivers Killed",            "driver_killed"),
    ("Drivers Grievous Injury",   "driver_grievous_injury"),
    ("Drivers Minor Injury",      "driver_minor_injury"),
    # Passenger
    ("Passengers Killed",         "passenger_killed"),
    ("Passengers Grievous Injury","passenger_grievous_injury"),
    ("Passengers Minor Injury",   "passenger_minor_injury"),
    # Pedestrian
    ("Pedestrians Killed",        "pedestrian_killed"),
    ("Pedestrians Grievous Injury","pedestrian_grievous_injury"),
    ("Pedestrians Minor Injury",  "pedestrian_minor_injury"),
    # Collision
    ("Collision Type",            "type_of_collision"),
    ("Collision Nature",          "collision_feature"),
    # Conditions
    ("Weather Condition",         "weather_condition"),
    ("Light Condition",           "light_condition"),
    ("Visibility",                "visibility"),
    ("Traffic Violation",         "traffic_violation"),
]

# Pre-compute parallel lists for fast row construction and header writing
HEADERS = [col[0] for col in EXPORT_COLUMNS]
FIELDS  = [col[1] for col in EXPORT_COLUMNS]


def row_values(accident, bs_id: int) -> list:
    """
    Extract and format an ORM accident object into a standardized list of row values.
    
    The first column is designated as the blackspot number (`bs_id`), which is 
    injected at runtime since it represents a computed grouping rather than a 
    direct attribute of the Accident model.
    
    Parameters
    ----------
    accident : object
        The SQLAlchemy ORM accident model instance.
    bs_id : int
        The computed Blackspot ID associated with this accident.

    Returns
    -------
    list
        An ordered list of formatted cell values corresponding to EXPORT_COLUMNS.
    """
    values = []
    for header, field in EXPORT_COLUMNS:
        # Inject the Blackspot ID dynamically
        if field == "__bs_id__":
            values.append(bs_id)
            continue
            
        # Dynamically fetch the attribute; default to None if missing
        raw = getattr(accident, field, None)
        
        # Format the output based on data type to ensure spreadsheet compatibility
        if raw is None:
            values.append("")
        elif isinstance(raw, datetime):
            # Format datetime to a clean, human-readable format (e.g., 15-Aug-2023 02:30 PM)
            values.append(raw.strftime("%d-%b-%Y %I:%M %p"))
        elif isinstance(raw, float):
            # Round high-precision floats (like GIS coordinates) to 6 decimal places
            values.append(round(raw, 6))
        else:
            # Fallback for text: sanitize using shared text utilities.
            # We strip out literal "Unknown" strings to leave cells blank for cleaner pivoting.
            cleaned = safe_text(str(raw))
            values.append("" if cleaned == "Unknown" else cleaned)
            
    return values


def build_accident_csv(accidents_with_bs: list[tuple[int, object]]) -> str:
    """
    Generate a raw CSV string from a list of accident records.

    Parameters
    ----------
    accidents_with_bs : list[tuple[int, object]]
        A list of tuples, where each tuple contains a Blackspot ID and an 
        Accident ORM object.

    Returns
    -------
    str
        The complete CSV formatted data as a string.
    """
    # Use StringIO to build the file in memory rather than hitting the disk
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write the column names first
    writer.writerow(HEADERS)
    
    # Convert and write each accident as a row
    for bs_id, acc in accidents_with_bs:
        writer.writerow(row_values(acc, bs_id))
        
    # Reset the buffer pointer to the beginning before reading out the string
    output.seek(0)
    return output.getvalue()


def build_accident_excel(
    accidents_with_bs: list[tuple[int, object]],
    meta_rows: list[tuple[str, object]],
    charts_dict: dict = None,
) -> io.BytesIO:
    """
    Generate a fully styled, multi-sheet Excel workbook from accident records.
    
    Includes professional formatting like frozen panes, auto-filters, alternating 
    row colors, a dedicated metadata/summary sheet, and optionally injects plotted 
    charts into a visualizations sheet.
    
    Parameters
    ----------
    accidents_with_bs : list[tuple[int, object]]
        List of (Blackspot ID, Accident ORM object) tuples for the main data sheet.
    meta_rows : list[tuple[str, object]]
        List of (label, value) tuples for the Summary metadata sheet.
    charts_dict : dict, optional
        Dictionary mapping specific chart names to their image byte buffers.

    Returns
    -------
    io.BytesIO
        An in-memory bytes buffer containing the generated .xlsx file, ready 
        to be returned as a streaming HTTP response.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accident Data"

    # Define reusable cell styles
    header_fill  = PatternFill("solid", fgColor="1E3A8A")  # Deep blue background
    header_font  = Font(bold=True, color="FFFFFF", size=10) # White bold text
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side    = Side(style="thin", color="D1D5DB")
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)
    alt_fill     = PatternFill("solid", fgColor="F8FAFC")  # Light gray alternating rows

    # Write and style the header row
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[1].height = 36

    # Write data rows, applying borders and zebra-striping
    for row_idx, (bs_id, acc) in enumerate(accidents_with_bs, start=2):
        ws.append(row_values(acc, bs_id))
        for cell in ws[row_idx]:
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Dynamically adjust column widths for readability
    WIDTHS = {
        "Blackspot #": 12, "Accident ID": 18, "District": 14,
        "Police Station": 20, "Accident Date Time": 22,
        "Latitude": 12, "Longitude": 12, "Road Name": 22,
        "Road Classification": 20, "Severity": 18, "No of Vehicles": 12,
    }
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = WIDTHS.get(header, 14)

    # Freeze the top header row and enable auto-filtering on the data
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Build Summary Sheet ---
    meta = wb.create_sheet("Summary")
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 36
    bold = Font(bold=True, size=10)
    for r, (label, value) in enumerate(meta_rows, start=1):
        meta.cell(r, 1, label).font = bold
        meta.cell(r, 2, str(value))

    # --- Build Visualizations Sheet (if charts provided) ---
    buf = io.BytesIO()
    
    if charts_dict:
        try:
            # openpyxl.drawing.image depends on 'Pillow'. We wrap in try/except 
            # so the export doesn't crash if the environment lacks image support.
            from openpyxl.drawing.image import Image as OpenpyxlImage
            viz_sheet = wb.create_sheet("Visualizations")
            
            # Map known chart keys (usually generated by matplotlib/seaborn) to static Excel cell anchors
            positions = {
                "severity": "B2",
                "road_type": "K2",
                "collision_type": "T2",
                "vehicles_involved": "B25",
                "weather": "K25",
                "light": "T25",
                "collision_nature": "B48",
                "time_of_day": "K48",
                "monthly_seasonality": "B71",
                "annual_trend": "K71",
                "weekend_vs_weekday": "T71",
            }
            
            # Insert images into the mapped cells
            for key, chart_buf in charts_dict.items():
                if key in positions:
                    img = OpenpyxlImage(chart_buf)
                    viz_sheet.add_image(img, positions[key])
        except ImportError:
            pass # Pillow not installed; gracefully skip chart insertion

    # Save to buffer and rewind pointer
    wb.save(buf)
    buf.seek(0)
    return buf